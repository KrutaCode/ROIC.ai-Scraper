# For number storage and manipulation
import pandas as pd

# Excel writer utilities
import Excel_Writer.excel_writer_utils

# For excel access
import openpyxl as pyxl



class ExcelWriter(Excel_Writer.excel_writer_utils.ExcelWriterUtils):
    def __init__(self,ticker: str):
        self.ticker = ticker.upper()
        self.excel_file_name = f"{self.ticker}.xlsx"
        self.file_path = f"Your File Path\\ROIC\\Excel_Files\\{self.excel_file_name}"

        exists = self.check_if_file_exists()
        if exists:
            self.workbook = pyxl.load_workbook(self.file_path)
        elif not exists:
            print(f"\n[Excel File Not Found] - {self.excel_file_name} does not exist\n")
            self.create_new_excel_file()

        super().__init__(self.file_path)

    '-------------------------------------------------------'
    def create_new_excel_file(self):
        # If the file does not exist.
        self.workbook = pyxl.Workbook(self.excel_file_name)
        self.workbook.create_sheet("Summary")
        self.workbook.create_sheet("Income Statement")
        self.workbook.create_sheet("Balance Sheet")
        self.workbook.create_sheet("Cash Flow")
        self.workbook.save(self.file_path)
        print(f"\n[Excel File Created] - {self.excel_file_name} was created with these sheets: {self.workbook.sheetnames}\n")

    '-------------------------------------------------------'
    def write_to_file(self, summary_df: pd.DataFrame, income_statement_df: pd.DataFrame, balance_sheet_df: pd.DataFrame, cash_flow_df: pd.DataFrame):

        with pd.ExcelWriter(self.file_path) as writer:
            summary_df.to_excel(writer, "Summary")
            income_statement_df.to_excel(writer, "Income Statement")
            balance_sheet_df.to_excel(writer, "Balance Sheet")
            cash_flow_df.to_excel(writer, "Cash Flow")


