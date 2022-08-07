import os
import openpyxl as pyxl

class ExcelWriterUtils:
    def __init__(self,path):
        self.file_path = path

    '-------------------------------------------------------'
    def check_if_file_exists(self):
        return os.path.exists(self.file_path)

    '-------------------------------------------------------'
    def get_worksheet(self, sheet_name: str):
        wb = pyxl.load_workbook(self.file_path)
        worksheet = wb[sheet_name]
    '-------------------------------------------------------'
    def get_sheet_names(self):
        wb = pyxl.load_workbook(self.file_path)
        return wb.sheetnames
    '-------------------------------------------------------'
